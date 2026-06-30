import requests
from bs4 import BeautifulSoup
import json
import time
import sys
import io
import os
import tempfile
from urllib3.exceptions import InsecureRequestWarning

# 設定 stdout 為 UTF-8 編碼（解決 Windows 終端編碼問題）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# 關閉 SSL 驗證警告（因為 scrape.center SSL 憑證已過期）
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

BASE_URL = "https://ssr1.scrape.center"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


def fetch_page(url):
    """發送 HTTP GET 請求並回傳 BeautifulSoup 物件"""
    response = requests.get(url, headers=HEADERS, verify=False, timeout=15)
    response.encoding = "utf-8"
    if response.status_code != 200:
        raise Exception(f"HTTP {response.status_code}")
    return BeautifulSoup(response.text, "lxml")


def parse_index_page(soup):
    """從列表頁解析電影基本資訊"""
    movies = []
    # 每個電影卡片
    items = soup.select(".el-card.item")
    if not items:
        return movies

    for item in items:
        movie = {}

        # 電影名稱 (中文 - 英文)
        name_tag = item.select_one("h2.m-b-sm")
        if name_tag:
            full_name = name_tag.get_text(strip=True)
            # 分離中英文名稱，格式: "霸王別姬 - Farewell My Concubine"
            if " - " in full_name:
                cn_name, en_name = full_name.split(" - ", 1)
            else:
                cn_name, en_name = full_name, ""
            movie["name_cn"] = cn_name
            movie["name_en"] = en_name

        # 電影類型（類別）
        category_tags = item.select(".category span")
        movie["categories"] = [tag.get_text(strip=True) for tag in category_tags]

        # 國家 / 時長（第一個 .info）
        info_divs = item.select(".info")
        if len(info_divs) >= 1:
            spans = info_divs[0].select("span")
            if len(spans) >= 3:
                movie["country"] = spans[0].get_text(strip=True)
                movie["duration"] = spans[2].get_text(strip=True)
            elif len(spans) >= 1:
                movie["country"] = spans[0].get_text(strip=True)
                movie["duration"] = ""

        # 上映日期（第二個 .info）
        if len(info_divs) >= 2:
            date_spans = info_divs[1].select("span")
            movie["release_date"] = date_spans[0].get_text(strip=True) if date_spans else ""

        # 評分
        score_tag = item.select_one(".score")
        movie["score"] = score_tag.get_text(strip=True) if score_tag else ""

        # 封面圖片
        cover_tag = item.select_one("img.cover")
        movie["cover_url"] = cover_tag["src"] if cover_tag else ""

        # 詳情頁連結 & ID
        link_tag = item.select_one("a[href^='/detail/']")
        if link_tag:
            detail_path = link_tag["href"]
            movie["detail_url"] = BASE_URL + detail_path
            movie["id"] = detail_path.split("/")[-1]

        movies.append(movie)

    return movies


def parse_detail_page(soup):
    """從詳情頁解析更多資訊（劇情簡介等）"""
    detail = {}

    # 劇情簡介
    drama_tag = soup.select_one(".drama p")
    if drama_tag:
        detail["synopsis"] = drama_tag.get_text(strip=True)
    else:
        detail["synopsis"] = ""

    # 劇照（最多取前 5 張）
    photo_tags = soup.select(".photo .el-image__inner")
    detail["stills"] = [tag["src"] for tag in photo_tags[:5]]

    return detail


def get_total_pages():
    """從首頁取得總頁數"""
    try:
        soup = fetch_page(BASE_URL + "/")
        # 從分頁元件中找最後一頁的連結
        page_links = soup.select(".el-pager li.number a")
        if page_links:
            last_page = max(
                int(link["href"].split("/")[-1])
                for link in page_links
            )
            return last_page
    except Exception:
        pass
    return 1


def scrape_all_movies(fetch_detail=False):
    """爬取所有電影資料"""
    all_movies = []

    total_pages = get_total_pages()
    print(f"開始爬取 {BASE_URL}/ ...")
    print(f"偵測到總頁數: {total_pages}")

    for page in range(1, total_pages + 1):
        if page == 1:
            url = BASE_URL + "/"
        else:
            url = f"{BASE_URL}/page/{page}"

        print(f"\n[爬取] 第 {page}/{total_pages} 頁: {url}")
        try:
            soup = fetch_page(url)
            movies = parse_index_page(soup)

            if fetch_detail:
                for movie in movies:
                    print(f"  [詳情] 電影 {movie.get('id')} - {movie.get('name_cn')}")
                    try:
                        detail_soup = fetch_page(movie["detail_url"])
                        detail = parse_detail_page(detail_soup)
                        movie.update(detail)
                        time.sleep(0.5)
                    except Exception as e:
                        print(f"    [!] 詳情頁爬取失敗: {e}")

            all_movies.extend(movies)
            print(f"  取得 {len(movies)} 部電影")

            time.sleep(1)
        except Exception as e:
            print(f"  [X] 頁面爬取失敗: {e}")

    return all_movies


def save_to_json(data, filename="movies.json"):
    """將資料儲存為 JSON 檔案"""
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n>>> 資料已儲存至 {filename}，共 {len(data)} 筆")


def save_to_csv(data, filename="movies.csv"):
    """將資料儲存為 CSV 表格（UTF-8 with BOM，Excel 可直接開啟）"""
    import csv
    # 定義欄位順序
    fieldnames = ["id", "name_cn", "name_en", "categories", "country", "duration",
                  "release_date", "score", "cover_url", "detail_url"]
    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            # 將 categories 列表轉成逗號分隔字串
            row_copy = dict(row)
            if isinstance(row_copy.get("categories"), list):
                row_copy["categories"] = "、".join(row_copy["categories"])
            writer.writerow({k: row_copy.get(k, "") for k in fieldnames})
    print(f">>> 表格已儲存至 {filename}，共 {len(data)} 筆，可直接用 Excel 開啟")


def download_posters(data, folder="posters"):
    """下載所有海報圖片到本地資料夾"""
    os.makedirs(folder, exist_ok=True)
    total = len(data)
    print(f"\n正在下載海報圖片到 {folder}/ ...")

    for i, movie in enumerate(data):
        cover_url = movie.get("cover_url", "")
        if not cover_url:
            continue

        # 檔名: ID_電影名稱.jpg
        safe_name = movie.get("name_cn", f"movie_{movie.get('id', i)}")
        safe_name = "".join(c for c in safe_name if c not in r'\/:*?"<>|')
        ext = ".png" if ".png" in cover_url else ".jpg"
        movie_id = int(movie.get("id", i))
        filename = f"{movie_id:03d}_{safe_name}{ext}"
        filepath = os.path.join(folder, filename)

        try:
            resp = requests.get(cover_url, headers=HEADERS, verify=False, timeout=30)
            if resp.status_code == 200:
                with open(filepath, "wb") as f:
                    f.write(resp.content)
        except Exception as e:
            print(f"  [!] 海報下載失敗 ID={movie.get('id')}: {e}")

        if (i + 1) % 20 == 0:
            print(f"  已下載 {i + 1}/{total} ...")

    downloaded = len(os.listdir(folder))
    print(f">>> 海報已儲存至 {folder}/ 資料夾，共 {downloaded} 張")


def save_to_excel(data, filename="movies.xlsx"):
    """下載海報圖片並寫入 Excel 表格（含嵌入式海報）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    wb = Workbook()
    ws = wb.active
    ws.title = "電影列表"

    # 欄位定義: (標題, key, 欄寬)
    columns = [
        ("ID", "id", 6),
        ("海報", None, 18),
        ("電影名稱", "name_cn", 16),
        ("英文名稱", "name_en", 30),
        ("類別", "categories", 22),
        ("國家", "country", 16),
        ("片長", "duration", 10),
        ("上映日期", "release_date", 16),
        ("評分", "score", 8),
        ("詳情連結", "detail_url", 40),
    ]

    # 樣式
    header_font = Font(name="微軟正黑體", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 寫入表頭
    for col_idx, (title, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # 暫存目錄放海報
    poster_dir = tempfile.mkdtemp(prefix="posters_")
    posters = {}

    print(f"\n正在下載海報圖片...")
    total = len(data)
    for i, movie in enumerate(data):
        cover_url = movie.get("cover_url", "")
        poster_path = None

        # 下載海報
        if cover_url:
            ext = ".jpg"
            if ".png" in cover_url:
                ext = ".png"
            poster_path = os.path.join(poster_dir, f"poster_{movie.get('id', i)}{ext}")
            try:
                resp = requests.get(cover_url, headers=HEADERS, verify=False, timeout=30)
                if resp.status_code == 200:
                    with open(poster_path, "wb") as f:
                        f.write(resp.content)
                    # 統一只留 200px 高
                    img = PILImage.open(poster_path)
                    ratio = 200 / img.height
                    new_w = int(img.width * ratio)
                    img = img.resize((new_w, 200), PILImage.LANCZOS)
                    img.save(poster_path)
                    posters[i] = poster_path
                else:
                    poster_path = None
            except Exception:
                poster_path = None

        if (i + 1) % 20 == 0:
            print(f"  已下載 {i + 1}/{total} ...")

    print(f"  海報下載完成，開始寫入 Excel ...")

    # 寫入資料列
    for i, movie in enumerate(data):
        row = i + 2
        cats = "、".join(movie.get("categories", [])) if isinstance(movie.get("categories"), list) else ""

        values = [
            int(movie.get("id", 0)),
            None,  # 海報稍後用圖片嵌入
            movie.get("name_cn", ""),
            movie.get("name_en", ""),
            cats,
            movie.get("country", ""),
            movie.get("duration", ""),
            movie.get("release_date", ""),
            float(movie.get("score", 0)),
            movie.get("detail_url", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            if col_idx == 2:
                continue  # 跳過海報文字
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.border = thin_border
            if col_idx in (1, 9):  # ID、評分置中
                cell.alignment = center_align
            else:
                cell.alignment = cell_align

        # 嵌入海報圖片
        if i in posters:
            try:
                img = XLImage(posters[i])
                img.width = 133   # 等比縮放至 200px 高
                img.height = 200
                cell = ws.cell(row=row, column=2)
                ws.add_image(img, cell.coordinate)
            except Exception:
                pass

        # 設定列高
        ws.row_dimensions[row].height = 205

    # 凍結首列
    ws.freeze_panes = "A2"

    # 加入第二個工作表：統計摘要
    ws2 = wb.create_sheet("統計摘要")
    scores = [float(m["score"]) for m in data if m.get("score")]
    all_cats = {}
    for m in data:
        for cat in m.get("categories", []):
            all_cats[cat] = all_cats.get(cat, 0) + 1

    summary_data = [
        ("項目", "數值"),
        ("總電影數", len(data)),
        ("最高評分", max(scores) if scores else "-"),
        ("最低評分", min(scores) if scores else "-"),
        ("平均評分", f"{sum(scores)/len(scores):.2f}" if scores else "-"),
    ]
    for r, (k, v) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=r, column=1, value=k)
        c2 = ws2.cell(row=r, column=2, value=v)
        if r == 1:
            c1.font = header_font; c1.fill = header_fill; c1.alignment = header_align
            c2.font = header_font; c2.fill = header_fill; c2.alignment = header_align
        c1.border = thin_border; c2.border = thin_border

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14

    # 類別分布
    row_start = len(summary_data) + 2
    ws2.cell(row=row_start, column=1, value="類別分布").font = Font(bold=True, size=12)
    row_start += 1
    for cat, cnt in sorted(all_cats.items(), key=lambda x: -x[1]):
        ws2.cell(row=row_start, column=1, value=cat).border = thin_border
        ws2.cell(row=row_start, column=2, value=cnt).border = thin_border
        row_start += 1

    wb.save(filename)
    print(f">>> Excel 已儲存至 {filename}（含海報圖片），共 {len(data)} 筆")

    # 清理暫存海報
    for p in posters.values():
        try:
            os.remove(p)
        except Exception:
            pass
    try:
        os.rmdir(poster_dir)
    except Exception:
        pass


def print_table(movies, limit=20):
    """印出電影資料表格"""
    # 欄位與寬度設定
    cols = [
        ("ID", "id", 3),
        ("電影名稱", "name_cn", 14),
        ("英文名稱", "name_en", 28),
        ("類別", "categories", 16),
        ("國家", "country", 12),
        ("片長", "duration", 8),
        ("上映日期", "release_date", 14),
        ("評分", "score", 5),
    ]

    # 表頭
    header = "│ " + " │ ".join(f"{c[0]:^{c[2]}}" for c in cols) + " │"
    sep = "├" + "┼".join("─" * (c[2] + 2) for c in cols) + "┤"
    top = "┌" + "┬".join("─" * (c[2] + 2) for c in cols) + "┐"
    bot = "└" + "┴".join("─" * (c[2] + 2) for c in cols) + "┘"

    print("\n" + top)
    print(header)
    print(sep)

    count = 0
    for m in movies:
        if count >= limit:
            break
        cats = "、".join(m.get("categories", [])) if isinstance(m.get("categories"), list) else m.get("categories", "")
        values = [
            m.get("id", ""),
            m.get("name_cn", ""),
            m.get("name_en", ""),
            cats,
            m.get("country", ""),
            m.get("duration", ""),
            m.get("release_date", ""),
            m.get("score", ""),
        ]
        # 截斷過長內容
        truncated = []
        for i, v in enumerate(values):
            s = str(v)
            w = cols[i][2]
            if len(s) > w:
                s = s[:w - 1] + "…"
            truncated.append(f"{s:<{w}}")
        print("│ " + " │ ".join(truncated) + " │")
        count += 1

    print(bot)
    if len(movies) > limit:
        print(f"(僅顯示前 {limit} 筆，共 {len(movies)} 筆。完整資料請見 movies.csv)")


if __name__ == "__main__":
    # ============================================================
    #  初學者模式：只爬列表頁資訊就好
    #  如需詳情頁（劇情簡介 + 劇照），將 fetch_detail 設為 True
    # ============================================================

    # 快速模式：只爬列表頁（速度快，資料已足夠）
    movies = scrape_all_movies(fetch_detail=False)

    # 完整模式：也爬詳情頁（較慢，資料更豐富）
    # movies = scrape_all_movies(fetch_detail=True)

    # 輸出表格
    print_table(movies, limit=20)

    # 下載海報圖片到本地
    download_posters(movies)

    # 儲存檔案
    save_to_excel(movies, "movies.xlsx")
    save_to_csv(movies, "movies.csv")
    save_to_json(movies, "movies.json")
